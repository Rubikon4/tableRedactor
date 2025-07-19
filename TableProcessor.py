import pandas as pd
from openpyxl import load_workbook
from datetime import datetime

class TableProcessor:
    def __init__(self, filepath):
        self.filepath = filepath
        self.df = None
        self.new_df = None
        self.date_value = None
        # self.is_Date = None
    def load(self, header=None):
        """Загружает Excel-файл в DataFrame без заголовков"""
        try:
            self.df = pd.read_excel(self.filepath, engine='openpyxl', header=None)
            self.new_df = self.df.copy() # для расширенной обработки
            return True
        except Exception as e:
            print(f"Ошибка загрузки файла: {e}")
            return False
    def clean_empty_rows(self):
        """Удаляет пустые строки"""
        if self.df is not None:
            self.df.dropna(how='all', inplace=True)
            # self.first_row_headTimerRunnerer()
    def filter_rows(self, start: int, end: int):
        """Оставляет строки от start до end включительно"""
        if self.df is not None:
            header_row = self.df.iloc[start - 1]
            data_rows = self.df.iloc[start:end]
            self.df = data_rows.copy()
            self.df.columns = header_row
            self.df.reset_index(drop=True, inplace=True)
    def filter_columns(self, start: int, end: int):
        """Оставляет столбцы от start до end включительно"""
        if self.df is not None:
            self.df = self.df.iloc[:, start:end + 1]
            self.first_row_header()
    def first_row_header(self):
        """Находит первую строку с данными и делает её заголовком"""
        for idx, row in self.df.iterrows():
            if row.notna().any():
                self.df.columns = row
                self.df = self.df.iloc[idx + 1:].reset_index(drop=True)
                break
    def export(self, output_path: str, format: str = 'xlsx'):
        """Сохраняет файл в указанный путь в формате xlsx или csv"""
        if self.df is not None:
            try:
                if format == 'csv':
                    self.df.to_csv(output_path, index=False)
                elif format == 'xlsx':
                    self.df.to_excel(output_path, index=False, engine='openpyxl')
                else:
                    raise ValueError("Неподдерживаемый формат")
                return True
            except Exception as e:
                print(f"Ошибка при сохранении: {e}")
                return False
        return False

    """    Для Master    """

    def clean_empty_rows_master(self):
        """Удаляет пустые строки"""
        if self.df is not None:
            self.new_df.dropna(how='all', inplace=True)
            # self.first_row_headTimerRunnerer()
    def set_header_by_number(self, row_number):
        """Устанавливает заголовки из строки с номером row_number"""
        idx = int(row_number) - 1
        if self.df is not None and idx < len(self.df):
            self.new_df.columns = self.df.iloc[idx]
            self.new_df = self.df.iloc[idx + 1:].reset_index(drop=True)
    def select_rows(self, rows_range):
        """Выбирает строки с данными от start_row до end_row по их номеру"""
        if len(rows_range.split()) > 1:
            start, end = map(int, rows_range.split())
            start_idx = start - 1
            if self.df is None or start_idx >= len(self.df):
                return
            end_idx = end
        else:
            start = int(rows_range.split()[0])
            start_idx = start - 1
            for i in range(start_idx, len(self.df)):
                if self.df.iloc[i].isnull().all():
                    end_idx = i
                    break
            else:
                end_idx = len(self.df)
        self.new_df = self.df.iloc[start_idx:end_idx].reset_index(drop=True) 

    def extract_date(self, mode, entry):
        """Обрабатывает дату по выбранному режиму"""
        if mode == "empty" and entry == None:
            return
        elif mode == "auto" and entry == None:
            for row in self.df.itertuples(index=False):
                for cell in row:
                    result = self.parse_date(cell)  
                    if result:  # добавить обработку ошибки
                        self.date_value = result
        elif mode == "cell" and entry != None:
            self.date_value = self.parse_date(self.extract_date_from_cell(entry))
                
    def parse_date(self, val):
        """Приводит дату к строке в формате дд.мм.гггг"""
        if isinstance(val, datetime):
            return val.strftime("%d.%m.%Y")
        elif isinstance(val, str):
            try:
                dt = pd.to_datetime(val, dayfirst=True)
                return dt.strftime("%d.%m.%Y")
            except:
                return None
        return None 
          
    def extract_date_from_cell(self, cell_ref):
        """Возвращает дату из указанной ячейки"""
        wb = load_workbook(self.filepath, read_only=True)
        ws = wb.active
        return ws[cell_ref].value

    def add_date_column(self):
        """Добавляет колонку 'Дата' ко всем строкам, если дата уже считана"""
        if self.date_value != None:
            if self.df is not None and getattr(self, "date_value", None):
                self.new_df["Дата"] = self.date_value
        else:
            return

    def export_master(self, output_path: str, format: str = 'xlsx'):
        """Сохраняет файл в указанный путь в формате xlsx или csv"""
        if self.new_df is not None:
            try:
                if format == 'csv':
                    self.df.to_csv(output_path, index=False)
                elif format == 'xlsx':
                    self.df.to_excel(output_path, index=False, engine='openpyxl')
                else:
                    raise ValueError("Неподдерживаемый формат")
                return True
            except Exception as e:
                print(f"Ошибка при сохранении: {e}")
                return False
        return False


        





    def extract_date_from_column(self, col_letter):
        """Считывает дату из первого непустого значения в колонке"""
        col_index = pd.ExcelFile(self.filepath).parse(header=None).columns.get_loc(col_letter)
        col = self.df.iloc[:, col_index].dropna()
        if not col.empty:
            self.date_value = self.parse_date(col.iloc[0])



    def export_xls_with_date_name(self, output_folder, template_name):
        """Сохраняет в .xlsx с добавлением даты в имя файла"""
        filename = template_name
        if getattr(self, "date_value", None):
            filename += f"_{self.date_value}"
        path = f"{output_folder}/{filename}.xlsx"
        self.df.to_excel(path, index=False, engine="openpyxl")